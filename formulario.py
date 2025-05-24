import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

# Nombre del archivo Excel
EXCEL_FILE = "datos_usuarios.xlsx"

# Crear archivo si no existe
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(["Nombre", "Apellido", "Edad", "Email", "Telefono"])  # encabezados
    wb.save(EXCEL_FILE)

# Función para guardar datos
def guardar_datos():
    nombre = entry_nombre.get()
    apellido = entry_apellido.get()
    edad = entry_edad.get()
    email = entry_email.get()
    telefono = entry_telefono.get()

    if not nombre or not apellido or not edad or not email or not telefono:
        messagebox.showwarning("Campos vacíos", "Todos los campos son obligatorios.")
        return

    try:
        edad = int(edad)
    except ValueError:
        messagebox.showerror("Error de formato", "La edad debe ser un número.")
        return
    
    try:
        telefono = int(telefono)
    except ValueError:
        messagebox.showerror("Error de formato", "el telefono debe ser un número.")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Usuarios"]
    ws.append([nombre, apellido, edad, email, telefono])
    wb.save(EXCEL_FILE)

    messagebox.showinfo("Éxito", "Datos guardados correctamente.")
    limpiar_campos()

def limpiar_campos():
    entry_nombre.delete(0, tk.END)
    entry_apellido.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)

# Crear ventana
ventana = tk.Tk()
ventana.title("Formulario a Excel")
ventana.geometry("450x350")
ventana.resizable(False, False)

# Etiquetas y entradas
tk.Label(ventana, text="Nombre:").pack(pady=5)
entry_nombre = tk.Entry(ventana, width=30)
entry_nombre.pack()

tk.Label(ventana, text="Apellido:").pack(pady=5)
entry_apellido = tk.Entry(ventana, width=30)
entry_apellido.pack()

tk.Label(ventana, text="Edad:").pack(pady=5)
entry_edad = tk.Entry(ventana, width=30)
entry_edad.pack()

tk.Label(ventana, text="Email:").pack(pady=5)
entry_email = tk.Entry(ventana, width=30)
entry_email.pack()

tk.Label(ventana, text="Telefono:").pack(pady=5)
entry_telefono = tk.Entry(ventana, width=30)
entry_telefono.pack()

# Botones
tk.Button(ventana, text="Guardar en Excel", command=guardar_datos).pack(pady=10)
tk.Button(ventana, text="Limpiar", command=limpiar_campos).pack()

# Ejecutar app
ventana.mainloop()
